
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

excel_file=Path("data")/"sales.xlsx"
df = pd.read_excel(excel_file,sheet_name="Sheet1")
print(df.head())



for index, row in df.iterrows():
    print(f"Row {index}: {row['employee']}, {row['sales']}, {row['quarter']}")



excel_formulas=load_workbook(excel_file,data_only=True)
sheet=excel_formulas.active
print(sheet["A1"].value) #* Expected outcome =SUM(A2:A5)


excel_values=load_workbook(excel_file,data_only=True)
sheet=excel_values.active

#!FORMATTTED CELLS
workbook=load_workbook(excel_file)
sheet=workbook.active
cell=sheet["A1"]
print(f"Cell Value: {cell.value}")
print(cell.font.name)
print(cell.font.size)
print(cell.font.bold)



